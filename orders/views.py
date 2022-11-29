from openpyxl import Workbook,load_workbook
from openpyxl.worksheet.table import Table
import csv
from django.db import connection
from django.http import HttpResponse
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.core.exceptions import PermissionDenied
from django.utils.decorators import classonlymethod
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.generics import RetrieveAPIView
from rest_framework.generics import ListAPIView
from rest_framework.generics import GenericAPIView
from rest_framework.mixins import CreateModelMixin, UpdateModelMixin, DestroyModelMixin
from rest_framework.viewsets import ViewSetMixin
from testproject.pagination import CustomPagination
from orders.models import Order
from orders.models import OrderItem
from orders.serializers import OrderSerializer
from users.auth import JWTAuthentication
from users.Signals import user_activity_signal


class OrderGenericAPIView(GenericAPIView,
                          PermissionRequiredMixin
                          ):
    def get_permission_required(self):
        if self.request.method == 'GET':
            return ['view_'+ self.model.lower()]
        if self.request.method == 'POST':
            return ['add_'+ self.model.lower()]
        if self.request.method == 'DELETE':
            return ['delete_'+ self.model.lower()]
        if self.request.method == 'PUT':
            return ['change_'+ self.model.lower()]

    def has_permission(self):
        return all(
            self.request.user.user_permissions.filter(codename=i).exists() for i in self.get_permission_required()
            )
    def get_object(self):
        print(self.get_permission_required(), self.request.user.user_permissions.all())
        if(self.has_permission()):
            print("user has permission")
            return super().get_object()
        else:
            print("no permission granted")
            raise PermissionDenied("good luck next time")

    authentication_classes = [JWTAuthentication]
    queryset = Order.objects.all()
    serializer_class = OrderSerializer


class GetAllOrder(OrderGenericAPIView, ListAPIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination


class GetanOrder(
    OrderGenericAPIView,
    RetrieveAPIView,
    CreateModelMixin,
    UpdateModelMixin,
    DestroyModelMixin,
    ViewSetMixin):

    # @classonlymethod
    # def as_view(self, cls, actions=None, **initkwargs):
    #     return super(ViewsetMixin, self).as_view(self, cls, actions,**initkwargs)
    
    # permission_classes = [IsAuthenticated]
    # permission_required = ('view_order','add_order','delete_order')
    
    def post(self, request, *args, **kwargs):
        if(self.has_permission()):
            user_activity_signal.send(sender=request.user, activity='have created a order')
            return self.create(request, *args, **kwargs)
        raise PermissionDenied

    def put(self, request, *args, **kwargs):
        user_activity_signal.send(sender=request.user, activity='have updated a order')
        return self.update(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        user_activity_signal.send(sender=request.user,activity='have deleted a order')
        return self.destroy(request, *args, **kwargs)
    lookup_field = "id"
    lookup_url_kwarg = "pk"


class ExportAPIView(APIView):
    # authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="order.csv"'
        orders = Order.objects.all()
        writer = csv.writer(response)
        writer.writerow(['ID', 'Name', 'Email', 'Product Title', 'Price', 'Quantity', '\n'])
        for order in orders:
            print("order", order)
            writer.writerow([order.id, order.name, order.email, '\n'])
            orderItems = OrderItem.objects.all().filter(order_id=order.id)
            print(orderItems)
            for item in orderItems:
                print("items", item)
                writer.writerow(["    ", item.product_title, item.price, item.quantity, '\n'])
        return response


class ChartAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, _):
        with connection.cursor() as cursor:
            cursor.execute("""
            SELECT DATE_FORMAT(o.created_at, '%Y-%m-%d') as date, sum(i.quantity * i.price) as sum
            FROM orders_order as o 
            JOIN orders_orderitem as i ON o.id = i.order_id GROUP BY date
            """)
            row = cursor.fetchall()

        data = [{
            'date': result[0],
            'sum': result[1]}
            for result in row]

        return Response({
            'data': data 
        })
class CreateOrder(OrderGenericAPIView,CreateModelMixin):
    def post(self, request, *args, **kwargs):
        wb = load_workbook(request.FILES['file'].file)
        ws = wb.active;
        l=[]
        # print(ws.tables)#,ws.tables[0].column_names)
        for name,rang in ws.tables.items():
            table_range : list = list(ws[rang]);
            columns_name = table_range.pop(0);
            for cells in table_range:
                try:
                    order = Order.objects.create(
                    email=cells[0].value,
                    first_name=cells[1].value,
                    last_name=cells[2].value)
                    l.append(OrderSerializer(order).data)
                except:
                    
                    continue
        return Response(
                l
                )